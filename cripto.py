import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime

class Leitura:

    def sheets(self, link, nome, j):
        wb = load_workbook(filename = 'criptomoedas.xlsx')
        wb.active = j
        ws= wb.active

        sheet_ranges = wb[nome]
        url = []
        vetor_bitcoin = []
        linha = str(5)
        contador = 0
        contador_repetido = 0
        posicao_bitcoin = 0
        i = 0

        while sheet_ranges['C'+linha].value != "FIM":

            nova_posicao = link+sheet_ranges['C'+linha].value+"/"
            
            if len(url) == 0:
                url.append(link+sheet_ranges['C'+linha].value+"/")
            else:
                for i in url:
                    if i == nova_posicao:
                        contador_repetido = contador_repetido + 1
                        if sheet_ranges['C'+linha].value == "bitcoin":
                            linha = int(linha)
                            vetor_bitcoin.append(linha)
                            linha = str(linha)

            if contador_repetido == 0 and linha != "5":
                url.append(link+sheet_ranges['C'+linha].value+"/")
                if sheet_ranges['C'+linha].value == "bitcoin":
                    linha = int(linha)
                    primeiro_bitcoin = linha
                    linha = str(linha)
                
            contador_repetido = 0
            linha = int(linha)
            linha = linha + 1
            linha = str(linha)


        linha = int(linha)
        ultima_linha = linha - 1
        contador = 0
        posicao_bitcoin = 0
        contador_altura = 5


        while contador_altura <=ultima_linha:

            for i in vetor_bitcoin:
                if contador_altura == i:
                    posicao_bitcoin = contador_altura

            if posicao_bitcoin>0:
                req = requests.get(url[primeiro_bitcoin-5])
                
            else:
                req = requests.get(url[contador])
                
            soup = BeautifulSoup(req.content, 'html.parser')
            valor_real = soup.find('div', class_='priceValue___11gHJ')
            valor_real = valor_real.next_element
            valor_real = valor_real.split("$")
            valor_real = valor_real[1]

            teste_array = list(valor_real)
            teste_array = teste_array[0]

            if len(valor_real) > 7 and teste_array!="0":
                valor_real = valor_real.split(",")
                valor_real = valor_real[0] + valor_real[1]

            valor_real = float(valor_real)
            concatenar = str(contador_altura)
            quantidade = float(sheet_ranges['D'+concatenar].value)
            ws['F'+concatenar] = valor_real*quantidade

            if posicao_bitcoin>0 or sheet_ranges['C'+concatenar].value == "bitcoin":
                ws['E'+concatenar] = float(quantidade)

            else:
                valor_bitcoin = soup.find('p', class_='sc-10nusm4-0 bspaAT')
                valor_bitcoin = valor_bitcoin.next_element
                valor_bitcoin = valor_bitcoin.split(" ")
                valor_bitcoin = valor_bitcoin[0]
                if valor_bitcoin == "<0.00000001":
                    valor_bitcoin = 0
                else:
                    valor_bitcoin = float(valor_bitcoin)
                ws['E'+concatenar] = valor_bitcoin*quantidade
        

            if posicao_bitcoin == 0:
                contador = contador + 1
            else:
                posicao_bitcoin = 0
            
            contador_altura = contador_altura + 1
        
        ws["B2"] = datetime.now()

        if j == 2:
            wb.active = 0
            ws= wb.active

        wb.save('criptomoedas.xlsx')

        
link = 'https://coinmarketcap.com/pt-br/currencies/'

reader = Leitura()
nome = ['Victor', 'Daniela', 'Barbara']
reader.sheets(link, nome[0], 0)
reader.sheets(link, nome[1], 1)
reader.sheets(link, nome[2], 2)
